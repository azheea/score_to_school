import json
import openpyxl
from tqdm import tqdm
import threading
from concurrent.futures import ThreadPoolExecutor
import os
import Get_informations.GetScore as GetScore
import Get_informations.Get_province_id as Get_Province_Id
import Get_informations.Get_school_id as Get_School_Id



def main(province_id,want,year,thread_num):
    path = os.getcwd()  # 获取当前工作目录的路径
    province_id = str(province_id)  # 将province_id转换为字符串类型
    is_hugescratch = True
    province_id = Get_Province_Id.get_province_id(province_id)  # 获取省份ID
    all_school_id = Get_School_Id.get_school_ids()  # 获取学校名称和对应的ID字典

    def fetch_score(line, worksheet, all_school_id, province_id, year, want, is_hugescratch):
        school_name = worksheet.cell(row=line + 1, column=4).value  # 获取工作表中指定单元格的值，即学校名称
        school_id = all_school_id.get(school_name)  # 根据学校名称从学校名称和对应的ID字典中获取学校ID
        if school_id:  # 如果学校ID存在
            try:
                worksheet.cell(row=line + 1, column=3).value = school_id  # 将学校ID写入工作表中的指定单元格
                all_score = GetScore.getScore(schoolId=school_id, provinceId=province_id, year=year)  # 调用GetScore模块的getScore函数获取分数信息

                result = {}  # 创建空字典，用于存储结果
                for item in all_score["item"]:  # 遍历分数信息列表
                    if (not is_hugescratch and item['spname'] == want) or (is_hugescratch and want in item['spname']):  # 如果不是模糊搜索且专业名称与所选专业相同，或者是模糊搜索且所选专业包含在专业名称中
                        result['spname'] = item['spname']  # 存储专业名称
                        result['min'] = item['min']  # 存储最低分
                        result['min_section'] = item['min_section']  # 存储最低分段
                        break  # 跳出循环 

                if result.get("min") is not None and result.get("min_section") is not None:  # 如果最低分和最低分段存在
                    with lock:  # 获取锁
                        worksheet.cell(row=line + 1, column=9).value = result.get("min")  # 将最低分写入工作表中的指定单元格
                        worksheet.cell(row=line + 1, column=10).value = result.get("min_section")  # 将最低分段写入工作表中的指定单元格
                        worksheet.cell(row=line + 1, column=11).value = result.get("spname")  # 将专业名称写入工作表中的指定单元格
                        # print(result)
            except Exception as e:  # 捕获异常并将异常信息存储到变量e中
                error = str(e)  # 将异常信息转换为字符串类型
                errors.append(f"{school_name},{error}")  # 将学校名称和异常信息添加到错误列表中
                

    workbook = openpyxl.load_workbook(f'{path}\schools.xlsx')  # 加载Excel文件
    worksheet = workbook['Sheet1']  # 获取工作表
    total_lines = worksheet.max_row - 1  # 获取工作表中的总行数
    lock = threading.Lock()  # 创建锁对象
    errors = []  # 创建空列表，用于存储错误信息

    with tqdm(total=total_lines, ncols=80, dynamic_ncols=True) as pbar:  # 使用tqdm创建进度条
        with ThreadPoolExecutor(max_workers=thread_num) as executor:  # 创建线程池
            for line in range(total_lines):  # 遍历总行数
                executor.submit(fetch_score, line, worksheet, all_school_id, province_id, year, want, is_hugescratch)  # 提交任务给线程池

        workbook.save(f"{path}\{year}_{want}_学校分数排行.xlsx")  # 保存Excel文件
        print(f"\n工作完成,结果已经输出在{path}\{year}_{want}_学校分数排行.xlsx")  # 打印完成信息及结果文件路径



if __name__ == "__main__":
    with open("config.json",encoding="utf-8") as f:
        configs = json.load(f)
        province_id = configs["province_id"]
        want = configs["want"]
        year = configs["year"]
        thread_num = configs["thread_num"]
    print("author 啊这.")
    print("仓库地址:\nhttps://github.com/azheea/score_to_school")

    main(province_id,want,year,thread_num)