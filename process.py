import json
import openpyxl
from tqdm import tqdm
import threading
from concurrent.futures import ThreadPoolExecutor
import os
import Get_informations.GetScore as GetScore
import Get_informations.Get_province_id as Get_Province_Id
import Get_informations.Get_school_id as Get_School_Id

# 调用分数获取逻辑
def process(province_id, want, year, thread_num, ethnic_minority,output,rank):
    province_id = str(province_id)  # 确保省份ID为字符串
    is_hugescratch = True  # 是否启用模糊搜索
    path = os.getcwd()  # 获取当前工作目录
    province_id = Get_Province_Id.get_province_id(province_id)  # 获取标准化的省份ID
    all_school_id = Get_School_Id.get_school_ids()  # 获取学校名称与ID的映射

    if output == "excel":
        # 分数获取函数，处理单行数据
        def fetch_score(line, worksheet, all_school_id, province_id, year, want, is_hugescratch,output):
            school_name = worksheet.cell(row=line + 1, column=4).value  # 从Excel中读取学校名称
            school_id = all_school_id.get(school_name)  # 根据学校名称获取学校ID
            if school_id:
                try:
                    worksheet.cell(row=line + 1, column=3).value = school_id  # 写入学校ID
                    all_score = GetScore.getScore(schoolId=school_id, provinceId=province_id, year=year)  # 获取分数数据

                    result = {}
                    # 遍历分数数据，筛选目标专业
                    for item in all_score["item"]:
                        if (not is_hugescratch and item['spname'] == want) or (is_hugescratch and want in item['spname']):
                            result['spname'] = item['spname']
                            result['min'] = item['min']
                            result['min_section'] = item['min_section']
                            break

                    # 如果找到目标专业，写入结果
                    if result.get("min") and result.get("min_section"):
                        with lock:
                            worksheet.cell(row=line + 1, column=9).value = result.get("min")
                            worksheet.cell(row=line + 1, column=10).value = result.get("min_section")
                            worksheet.cell(row=line + 1, column=11).value = result.get("spname")
                except Exception as e:
                    errors.append(f"{school_name},{str(e)}")  # 记录错误信息

        # 加载Excel文件
        workbook = openpyxl.load_workbook(f'{path}\schools.xlsx')
        worksheet = workbook['Sheet1']  # 选择工作表
        total_lines = worksheet.max_row - 1  # 计算总行数
        lock = threading.Lock()  # 创建线程锁
        errors = []  # 用于存储错误信息

        # 确保 all_school_id 是字典类型
        if not isinstance(all_school_id, dict):
            raise TypeError("all_school_id 应该是一个字典，但实际是: {}".format(type(all_school_id).__name__))

        # 使用多线程处理数据
        with tqdm(total=total_lines, ncols=80, dynamic_ncols=True) as pbar:
            with ThreadPoolExecutor(max_workers=thread_num) as executor:
                for line in range(total_lines):
                    executor.submit(fetch_score, line, worksheet, all_school_id, province_id, year, want, is_hugescratch)

            # 保存处理后的Excel文件
            workbook.save(f"{path}\{year}_{want}_学校分数排行.xlsx")
            print(f"\n工作完成,结果已经输出在{path}\{year}_{want}_学校分数排行.xlsx")
    
    #TODO: 添加json输出功能
    elif output == "json":
        # 修复 all_school_id 的迭代问题
        results = []
        for school_name, school_id in all_school_id.items():
            if "学院" in school_name or "职业" in school_name:
                continue
            try:
                result = []
                all_score = GetScore.getScore(schoolId=school_id, provinceId=province_id, year=year)  # 获取分数数据
                majors = all_score.get("item", [])
                for major in majors:
                    if (want in major["spname"] and ("专项"in major["spname"] or ethnic_minority == False)):
                        print(result)
                        result.append({major["spname"]: {"min": major["min"], "min_section": major["min_section"]}})
            except Exception as e:
                print(f"Error processing {school_name}: {e}")
            if result == []:
                continue
            results.append({school_name: result})

        # 保存结果到JSON文件
        output_file = f"{path}/{year}_{want}_学校分数排行.json"
        print(f"正在保存结果到 {output_file} ...")
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=4)
